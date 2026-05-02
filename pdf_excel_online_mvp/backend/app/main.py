from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pathlib import Path
from uuid import uuid4
from datetime import datetime
from typing import List
from io import BytesIO
import os
import json

import google.generativeai as genai
from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "storage" / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="PDF to Excel AI Enterprise")

genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


PROMPT_DDT_PAGE = """
Analizza questa pagina di un DDT scansionato.

Estrai SOLO i dati stampati.
Ignora scritte a mano, firme, timbri, segni di spunta, correzioni a penna e rumore.

Rispondi SOLO in JSON valido, senza markdown e senza testo extra.

Schema obbligatorio:

{
  "summary": {
    "tipo_documento": "DDT",
    "numero_documento": "",
    "data_documento": "",
    "mittente": "",
    "destinatario": "",
    "indirizzo_destinatario": "",
    "totale_pezzi": 0,
    "confidenza": 0
  },
  "righe": [
    {
      "codice": "",
      "descrizione": "",
      "ean": "",
      "quantita": 0,
      "confidenza": 0
    }
  ],
  "campi_da_verificare": [],
  "errori": []
}

Regole:
- leggi la tabella articoli riga per riga
- non inventare dati
- quantità sempre numerica
- data in formato GG/MM/AAAA
- se nella pagina non vedi testata o totale, lascia vuoto/0
- confidenza da 0 a 100
"""


def clean_json_text(text: str) -> str:
    return text.replace("```json", "").replace("```", "").strip()


def to_int(value, default=0):
    try:
        return int(str(value).replace(",", ".").split(".")[0])
    except Exception:
        return default


def normalize_page_data(data: dict) -> dict:
    summary = data.get("summary") or {}
    righe = data.get("righe") or []

    rows = []
    for row in righe:
        if not isinstance(row, dict):
            continue

        codice = str(row.get("codice") or "").strip()
        descrizione = str(row.get("descrizione") or "").strip()
        ean = str(row.get("ean") or "").strip()
        quantita = to_int(row.get("quantita"), 0)
        confidenza = to_int(row.get("confidenza"), 0)

        if not codice and not descrizione and not ean:
            continue

        rows.append({
            "codice": codice,
            "descrizione": descrizione,
            "ean": ean,
            "quantita": quantita,
            "confidenza": confidenza,
        })

    return {
        "summary": {
            "tipo_documento": summary.get("tipo_documento") or "DDT",
            "numero_documento": str(summary.get("numero_documento") or "").strip(),
            "data_documento": str(summary.get("data_documento") or "").strip(),
            "mittente": str(summary.get("mittente") or "").strip(),
            "destinatario": str(summary.get("destinatario") or "").strip(),
            "indirizzo_destinatario": str(summary.get("indirizzo_destinatario") or "").strip(),
            "totale_pezzi": to_int(summary.get("totale_pezzi"), 0),
            "confidenza": to_int(summary.get("confidenza"), 0),
        },
        "righe": rows,
        "campi_da_verificare": data.get("campi_da_verificare") or [],
        "errori": data.get("errori") or [],
    }


def merge_pages(page_results: list) -> dict:
    merged_summary = {
        "tipo_documento": "DDT",
        "numero_documento": "",
        "data_documento": "",
        "mittente": "",
        "destinatario": "",
        "indirizzo_destinatario": "",
        "totale_pezzi": 0,
        "confidenza": 0,
    }

    all_rows = []
    warnings = []
    errors = []

    for page_index, data in enumerate(page_results, start=1):
        summary = data.get("summary", {})

        for key in [
            "tipo_documento",
            "numero_documento",
            "data_documento",
            "mittente",
            "destinatario",
            "indirizzo_destinatario",
        ]:
            if not merged_summary.get(key) and summary.get(key):
                merged_summary[key] = summary.get(key)

        if summary.get("totale_pezzi"):
            merged_summary["totale_pezzi"] = summary.get("totale_pezzi")

        if summary.get("confidenza"):
            merged_summary["confidenza"] = max(
                merged_summary.get("confidenza", 0),
                summary.get("confidenza", 0)
            )

        for row in data.get("righe", []):
            row["pagina"] = page_index
            all_rows.append(row)

        for item in data.get("campi_da_verificare", []):
            warnings.append(f"Pagina {page_index}: {item}")

        for item in data.get("errori", []):
            errors.append(f"Pagina {page_index}: {item}")

    if not merged_summary["numero_documento"]:
        warnings.append("Numero documento mancante")

    if not merged_summary["data_documento"]:
        warnings.append("Data documento mancante")

    if not all_rows:
        warnings.append("Nessuna riga articolo rilevata")

    return {
        "summary": merged_summary,
        "righe": all_rows,
        "campi_da_verificare": warnings,
        "errori": errors,
    }


def split_pdf_pages(pdf_bytes: bytes) -> list:
    reader = PdfReader(BytesIO(pdf_bytes))
    pages = []

    for page in reader.pages:
        writer = PdfWriter()
        writer.add_page(page)

        buffer = BytesIO()
        writer.write(buffer)
        pages.append(buffer.getvalue())

    return pages


def build_excel_rows_only(result: dict, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dettaglio_Righe"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append([
        "File",
        "Pagina",
        "Numero Documento",
        "Data Documento",
        "Mittente",
        "Destinatario",
        "Codice",
        "Descrizione",
        "EAN",
        "Quantità",
        "Confidenza"
    ])

    summary = result.get("summary", {})
    righe = result.get("righe", [])

    for row in righe:
        ws.append([
            result.get("file", ""),
            row.get("pagina", ""),
            summary.get("numero_documento", ""),
            summary.get("data_documento", ""),
            summary.get("mittente", ""),
            summary.get("destinatario", ""),
            row.get("codice", ""),
            row.get("descrizione", ""),
            row.get("ean", ""),
            row.get("quantita", ""),
            row.get("confidenza", "")
        ])

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    wb.save(output_path)


async def process_single_pdf(file: UploadFile) -> dict:
    if not file.filename.lower().endswith(".pdf"):
        return {
            "file": file.filename,
            "status": "error",
            "error": "File non PDF"
        }

    contents = await file.read()

    try:
        pages = split_pdf_pages(contents)
    except Exception as e:
        return {
            "file": file.filename,
            "status": "error",
            "error": f"Errore lettura PDF: {str(e)}"
        }

    page_results = []
    model = genai.GenerativeModel("gemini-2.5-flash")

    for index, page_bytes in enumerate(pages, start=1):
        try:
            response = model.generate_content([
                {"mime_type": "application/pdf", "data": page_bytes},
                PROMPT_DDT_PAGE
            ])

            text = clean_json_text(response.text)

            try:
                raw_data = json.loads(text)
            except Exception:
                raw_data = {
                    "summary": {},
                    "righe": [],
                    "campi_da_verificare": [f"JSON non valido pagina {index}"],
                    "errori": [text],
                }

            page_results.append(normalize_page_data(raw_data))

        except Exception as e:
            page_results.append({
                "summary": {},
                "righe": [],
                "campi_da_verificare": [],
                "errori": [f"Errore AI pagina {index}: {str(e)}"],
            })

    merged = merge_pages(page_results)

    job_id = str(uuid4())
    output_path = OUTPUT_DIR / f"{job_id}.xlsx"

    result = {
        "file": file.filename,
        "data_elaborazione": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": merged.get("summary", {}),
        "righe": merged.get("righe", []),
        "campi_da_verificare": merged.get("campi_da_verificare", []),
        "errori": merged.get("errori", []),
        "pagine": len(pages),
    }

    build_excel_rows_only(result, output_path)

    return {
        "file": file.filename,
        "status": "completed",
        "job_id": job_id,
        "download_url": f"/api/download/{job_id}",
        "summary": result["summary"],
        "righe": result["righe"],
        "campi_da_verificare": result["campi_da_verificare"],
        "errori": result["errori"],
        "pagine": result["pagine"],
    }


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/api/convert")
async def convert(files: List[UploadFile] = File(...)):
    results = []

    for file in files:
        result = await process_single_pdf(file)
        results.append(result)

    return {
        "count": len(results),
        "documents": results
    }


@app.get("/api/download/{job_id}")
def download(job_id: str):
    xlsx_path = OUTPUT_DIR / f"{job_id}.xlsx"

    if not xlsx_path.exists():
        raise HTTPException(status_code=404, detail="File Excel non trovato.")

    return FileResponse(
        path=xlsx_path,
        filename=f"pdf_convertito_{job_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
